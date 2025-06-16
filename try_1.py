import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# 设置路径
main_file = '主表.xlsx'
sub_files = ['附表1.xlsx', '附表2.xlsx', '附表3.xlsx']  # 你可以动态读取文件夹
sheet_name = 'Sheet1'

# 读取所有表格
def read_and_clean_excel(file):
    df = pd.read_excel(file, sheet_name=sheet_name)
    if '负责人' in df.columns:
        df['负责人'] = df['负责人'].fillna(method='ffill')  # 向上填充负责人
    return df

# 读取主表
main_df = read_and_clean_excel(main_file)

# 合并附表
for file in sub_files:
    df = read_and_clean_excel(file)
    main_df = pd.concat([main_df, df], ignore_index=True)

# 合并同类项：按所有列去重（或者你只按部分列分组）
merged_df = main_df.drop_duplicates()

# 写入结果
wb = Workbook()
ws = wb.active
ws.title = "合并结果"

for r in dataframe_to_rows(merged_df, index=False, header=True):
    ws.append(r)

# 可选：自动合并“相邻重复的负责人单元格”
from openpyxl.utils import get_column_letter

def merge_repeated_cells(ws, col_name):
    col_index = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == col_name:
            col_index = idx + 1
            break
    if col_index is None:
        return

    start_row = 2
    current_value = ws.cell(row=start_row, column=col_index).value
    for row in range(3, ws.max_row + 2):
        val = ws.cell(row=row, column=col_index).value
        if val != current_value:
            if row - 1 > start_row:
                ws.merge_cells(start_row=start_row, start_column=col_index,
                               end_row=row - 1, end_column=col_index)
            start_row = row
            current_value = val

merge_repeated_cells(ws, '负责人')

# 保存结果
wb.save("合并结果.xlsx")
print("✅ 合并完成，已保存为合并结果.xlsx")
